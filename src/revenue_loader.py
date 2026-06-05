from __future__ import annotations

import calendar
import hashlib
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


FILE_PATTERN = re.compile(
    r"^Solutions_Revenue_(?P<month>[A-Za-z]+)_(?P<year>\d{4}).*\.xlsx$",
    re.IGNORECASE,
)

MONTH_NAME_TO_NUMBER = {
    name.lower(): index
    for index, name in enumerate(calendar.month_name)
    if name
}


@dataclass(frozen=True)
class AppConfig:
    data_folder: Path
    template_file: Path


CACHE_VERSION = "v3"

EXCLUDED_SHEET_KEYWORDS = (
    "summary",
    "by client",
    "combined",
    "practice",
    "service lines",
    "location",
    "new logos",
    "vlookup",
    "offset",
    "sheet1",
)


def _cache_root() -> Path:
    if os.name == "nt":
        base = Path(os.getenv("LOCALAPPDATA", str(Path.home() / "AppData" / "Local")))
    else:
        base = Path(os.getenv("XDG_CACHE_HOME", str(Path.home() / ".cache")))
    root = base / "revenue-analysis-dashboard"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _file_signature(path: Path) -> str:
    stats = path.stat()
    payload = f"{path.resolve()}|{stats.st_size}|{stats.st_mtime_ns}|{CACHE_VERSION}"
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def _cache_file_path(path: Path, namespace: str) -> Path:
    cache_dir = _cache_root() / namespace
    cache_dir.mkdir(parents=True, exist_ok=True)
    signature = _file_signature(path)
    stem = re.sub(r"[^A-Za-z0-9_-]", "_", path.stem)
    return cache_dir / f"{stem}_{signature}.pkl"


def _load_cached_frame(path: Path, namespace: str) -> pd.DataFrame | None:
    cache_file = _cache_file_path(path, namespace)
    if not cache_file.exists():
        return None
    try:
        frame = pd.read_pickle(cache_file)
        if isinstance(frame, pd.DataFrame):
            return frame
    except Exception:
        return None
    return None


def _save_cached_frame(path: Path, namespace: str, frame: pd.DataFrame) -> None:
    cache_file = _cache_file_path(path, namespace)
    frame.to_pickle(cache_file)


def _latest_cache_file(namespace: str) -> Path:
    cache_dir = _cache_root() / namespace
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir / "latest.pkl"


def _save_latest_cache(namespace: str, frame: pd.DataFrame) -> None:
    frame.to_pickle(_latest_cache_file(namespace))


def _load_latest_cache(namespace: str) -> pd.DataFrame | None:
    latest = _latest_cache_file(namespace)
    if not latest.exists():
        return None
    try:
        frame = pd.read_pickle(latest)
        if isinstance(frame, pd.DataFrame):
            return frame
    except Exception:
        return None
    return None


def _to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_date(value: object) -> pd.Timestamp | pd.NaT:
    if value is None or str(value).strip() == "":
        return pd.NaT
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    return parsed


def _engagement_label(raw_value: str) -> str:
    lowered = raw_value.lower()
    if "fixed" in lowered:
        return "FP"
    if "time" in lowered or "t&m" in lowered or "tm" in lowered:
        return "T&M"
    if lowered in {"fp", "t&m"}:
        return raw_value.upper() if lowered == "fp" else "T&M"
    return raw_value


def _detect_month_from_filename(path: Path) -> pd.Timestamp | None:
    match = FILE_PATTERN.match(path.name)
    if not match:
        return None
    month_name = match.group("month").lower()
    year = int(match.group("year"))
    month = MONTH_NAME_TO_NUMBER.get(month_name)
    if not month:
        return None
    return pd.Timestamp(year=year, month=month, day=1)


def _is_data_sheet(row9_values: list[object]) -> bool:
    text_values = [_to_string(v).lower() for v in row9_values]
    return "customer name" in text_values and "customer id" in text_values and "project name" in text_values


def _parse_revenue_workbook(path: Path) -> pd.DataFrame:
    month_ts = _detect_month_from_filename(path)
    if month_ts is None:
        return pd.DataFrame()

    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        lower_sheet = sheet_name.lower()
        if any(keyword in lower_sheet for keyword in EXCLUDED_SHEET_KEYWORDS):
            continue

        worksheet = workbook[sheet_name]
        row9 = list(next(worksheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=34, values_only=True), []))
        if not _is_data_sheet(row9):
            continue

        blank_streak = 0
        for row_values_raw in worksheet.iter_rows(min_row=10, min_col=1, max_col=27, values_only=True):
            row_values = list(row_values_raw)
            if sum(1 for value in row_values if value is None or str(value).strip() == "") >= 24:
                blank_streak += 1
                # Stop after repeated blank rows to avoid scanning large sheet tails.
                if blank_streak >= 20:
                    break
                continue
            blank_streak = 0

            customer_name = _to_string(row_values[0])
            customer_id = _to_string(row_values[1])
            project_id = _to_string(row_values[2])
            project_name = _to_string(row_values[3])

            is_summary_row = (
                not customer_name
                or not customer_id
                or customer_name.lower() in {" ", "total", "subtotal"}
                or project_name.lower() in {"", "total", "subtotal"}
            )

            if not is_summary_row and project_id:
                rows.append(
                    {
                        "source_file": path.name,
                        "month": month_ts,
                        "practice_sheet": sheet_name,
                        "customer_name": customer_name,
                        "customer_id": customer_id,
                        "project_id": project_id,
                        "project_name": project_name,
                        "engagement_model": _engagement_label(_to_string(row_values[6])),
                        "start_date": _normalize_date(row_values[7]),
                        "end_date": _normalize_date(row_values[8]),
                        "target_gm_pct": _to_number(row_values[10]) * (100 if _to_number(row_values[10]) <= 1 else 1),
                        "rev_month": _to_number(row_values[12]),
                        "cost_month": _to_number(row_values[13]),
                        "rev_ytd": _to_number(row_values[16]),
                        "cost_ytd": _to_number(row_values[17]),
                        "rev_jtd": _to_number(row_values[20]),
                        "cost_jtd": _to_number(row_values[21]),
                    }
                )

    workbook.close()

    if not rows:
        return pd.DataFrame()

    parsed = pd.DataFrame(rows)
    return _canonicalize_project_month(parsed)


def _canonicalize_project_month(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return data

    # Remove exact duplicates first, then aggregate to one row per project-month key.
    dedup = data.drop_duplicates(
        subset=[
            "month",
            "customer_id",
            "project_id",
            "project_name",
            "rev_month",
            "cost_month",
            "rev_ytd",
            "cost_ytd",
            "rev_jtd",
            "cost_jtd",
        ]
    ).copy()

    keys = ["month", "customer_id", "project_id"]
    grouped = (
        dedup.groupby(keys, dropna=False)
        .agg(
            source_file=("source_file", lambda s: " | ".join(sorted(set(str(v) for v in s if str(v).strip())))),
            practice_sheet=("practice_sheet", lambda s: " | ".join(sorted(set(str(v) for v in s if str(v).strip())))),
            customer_name=("customer_name", "first"),
            project_name=("project_name", "first"),
            engagement_model=("engagement_model", "first"),
            start_date=("start_date", "min"),
            end_date=("end_date", "max"),
            target_gm_pct=("target_gm_pct", "max"),
            rev_month=("rev_month", "sum"),
            cost_month=("cost_month", "sum"),
            rev_ytd=("rev_ytd", "sum"),
            cost_ytd=("cost_ytd", "sum"),
            rev_jtd=("rev_jtd", "sum"),
            cost_jtd=("cost_jtd", "sum"),
        )
        .reset_index()
    )

    grouped["margin_month"] = grouped["rev_month"] - grouped["cost_month"]
    grouped["gm_month_pct"] = _safe_gm_pct(grouped["margin_month"], grouped["rev_month"])
    grouped["margin_ytd"] = grouped["rev_ytd"] - grouped["cost_ytd"]
    grouped["gm_ytd_pct"] = _safe_gm_pct(grouped["margin_ytd"], grouped["rev_ytd"])
    grouped["margin_jtd"] = grouped["rev_jtd"] - grouped["cost_jtd"]
    grouped["gm_jtd_pct"] = _safe_gm_pct(grouped["margin_jtd"], grouped["rev_jtd"])

    return grouped


def _safe_gm_pct(margin: pd.Series, revenue: pd.Series) -> pd.Series:
    revenue_non_zero = revenue.where(revenue != 0)
    gm = (margin / revenue_non_zero) * 100
    return gm.fillna(0.0)


def _find_master_sheet(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    preferred = [s for s in workbook.sheetnames if "dashboard" in s.lower()]
    names_to_scan: Iterable[str] = preferred if preferred else workbook.sheetnames

    for sheet_name in names_to_scan:
        worksheet = workbook[sheet_name]
        for header_row in range(1, 25):
            headers = [_to_string(worksheet.cell(row=header_row, column=col).value).lower() for col in range(1, 40)]
            if "customer id" in headers and ("project id" in headers or "project code" in headers):
                return workbook, sheet_name, header_row

    return workbook, workbook.sheetnames[0], 1


def _combined_headers(worksheet, header_row: int, max_col: int = 60) -> list[str]:
    row_a = [_to_string(worksheet.cell(row=header_row, column=col).value) for col in range(1, max_col)]
    row_b = [_to_string(worksheet.cell(row=header_row + 1, column=col).value) for col in range(1, max_col)]
    combined = []
    for a, b in zip(row_a, row_b):
        combined.append(a if a else b)
    return combined


def load_master_projects(template_file: Path) -> pd.DataFrame:
    if not template_file.exists():
        return pd.DataFrame(columns=["customer_id", "project_id", "customer_name", "project_name"])

    cached = _load_cached_frame(template_file, "master")
    if cached is not None and not cached.empty:
        return cached

    try:
        workbook, sheet_name, header_row = _find_master_sheet(template_file)
    except PermissionError:
        # If template is locked, use the latest successful master cache if available.
        latest = _load_latest_cache("master")
        if latest is not None:
            return latest
        return pd.DataFrame(columns=["customer_id", "project_id", "customer_name", "project_name"])

    worksheet = workbook[sheet_name]

    def _resolve_header_map(headers: list[str]):
        lower = [h.lower() for h in headers]

        def _idx(*choices: str) -> int | None:
            for option in choices:
                if option in lower:
                    return lower.index(option)
            return None

        return {
            "c_id": _idx("customer id", "client id"),
            "c_name": _idx("customer name", "client name", "account name"),
            "p_id": _idx("project id", "project code"),
            "p_name": _idx("project name"),
            "e_model": _idx("t&m/fp", "engagement model"),
            "start_idx": _idx("start date"),
            "end_idx": _idx("end date"),
            "target_idx": _idx("target gm%", "target gm %"),
        }

    headers = [_to_string(worksheet.cell(row=header_row, column=col).value) for col in range(1, 60)]
    resolved = _resolve_header_map(headers)

    scan_row = header_row
    if resolved["c_id"] is None or resolved["p_id"] is None:
        for candidate in range(1, 60):
            candidate_headers = _combined_headers(worksheet, candidate, 60)
            candidate_resolved = _resolve_header_map(candidate_headers)
            if candidate_resolved["c_id"] is not None and candidate_resolved["p_id"] is not None:
                headers = candidate_headers
                resolved = candidate_resolved
                scan_row = candidate + 1
                break

    c_id = resolved["c_id"]
    c_name = resolved["c_name"]
    p_id = resolved["p_id"]
    p_name = resolved["p_name"]
    e_model = resolved["e_model"]
    start_idx = resolved["start_idx"]
    end_idx = resolved["end_idx"]
    target_idx = resolved["target_idx"]

    if c_id is None or p_id is None:
        latest = _load_latest_cache("master")
        if latest is not None:
            return latest
        return pd.DataFrame(columns=["customer_id", "project_id", "customer_name", "project_name"])

    rows: list[dict[str, object]] = []
    r = scan_row + 1
    while True:
        values = [worksheet.cell(row=r, column=col).value for col in range(1, 60)]
        if values.count(None) >= 56:
            if r > header_row + 10:
                break
        customer_id = _to_string(values[c_id])
        project_id = _to_string(values[p_id])
        if customer_id and project_id:
            target_raw = _to_number(values[target_idx]) if target_idx is not None else 0.0
            target_pct = target_raw * (100 if target_raw <= 1 else 1)
            rows.append(
                {
                    "customer_id": customer_id,
                    "project_id": project_id,
                    "customer_name": _to_string(values[c_name]) if c_name is not None else "",
                    "project_name": _to_string(values[p_name]) if p_name is not None else "",
                    "engagement_model": _engagement_label(_to_string(values[e_model])) if e_model is not None else "",
                    "start_date": _normalize_date(values[start_idx]) if start_idx is not None else pd.NaT,
                    "end_date": _normalize_date(values[end_idx]) if end_idx is not None else pd.NaT,
                    "target_gm_pct": target_pct,
                }
            )
        r += 1

    result = pd.DataFrame(rows)
    workbook.close()
    _save_cached_frame(template_file, "master", result)
    _save_latest_cache("master", result)
    return result


def load_revenue_files(data_folder: Path) -> pd.DataFrame:
    files = sorted(data_folder.glob("Solutions_Revenue_*.xlsx"))
    if not files:
        return pd.DataFrame()

    frames = []
    for path in files:
        cached = _load_cached_frame(path, "monthly")
        if cached is not None and not cached.empty:
            frames.append(cached)
            continue

        parsed = _parse_revenue_workbook(path)
        _save_cached_frame(path, "monthly", parsed)
        frames.append(parsed)

    frames = [f for f in frames if not f.empty]
    if not frames:
        return pd.DataFrame()

    data = pd.concat(frames, ignore_index=True)
    data["customer_id"] = data["customer_id"].astype(str)
    data["project_id"] = data["project_id"].astype(str)
    return data


def merge_with_master(revenue_data: pd.DataFrame, master_data: pd.DataFrame) -> pd.DataFrame:
    if revenue_data.empty and master_data.empty:
        return pd.DataFrame()

    if revenue_data.empty:
        result = master_data.copy()
        result["month"] = pd.NaT
        for col in [
            "rev_month",
            "cost_month",
            "margin_month",
            "gm_month_pct",
            "rev_ytd",
            "cost_ytd",
            "margin_ytd",
            "gm_ytd_pct",
            "rev_jtd",
            "cost_jtd",
            "margin_jtd",
            "gm_jtd_pct",
        ]:
            result[col] = 0.0
        return result

    if master_data.empty:
        # Master is mandatory: only accounts/projects listed in the template are allowed.
        return pd.DataFrame()

    master_data = master_data.copy()
    master_data["customer_id"] = master_data["customer_id"].astype(str)
    master_data["project_id"] = master_data["project_id"].astype(str)
    master_data = master_data.drop_duplicates(subset=["customer_id", "project_id"], keep="first")

    merged = master_data.merge(
        revenue_data,
        on=["customer_id", "project_id"],
        how="left",
        suffixes=("_master", ""),
    )

    for col in ["customer_name", "project_name", "engagement_model", "start_date", "end_date", "target_gm_pct"]:
        master_col = f"{col}_master"
        if master_col in merged.columns:
            merged[col] = merged[col].where(merged[col].notna() & (merged[col] != ""), merged[master_col])
            merged = merged.drop(columns=[master_col])

    numeric_cols = [
        "rev_month",
        "cost_month",
        "margin_month",
        "gm_month_pct",
        "rev_ytd",
        "cost_ytd",
        "margin_ytd",
        "gm_ytd_pct",
        "rev_jtd",
        "cost_jtd",
        "margin_jtd",
        "gm_jtd_pct",
    ]
    for col in numeric_cols:
        if col not in merged:
            merged[col] = 0.0
        merged[col] = merged[col].fillna(0.0)

    return merged


def build_project_snapshot(data: pd.DataFrame) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame()

    ordered = data.sort_values(["customer_name", "project_name", "month"], na_position="last")
    keys = ["customer_id", "project_id"]

    # Some master projects may not have monthly rows yet (month remains NaT in the group).
    # idxmax raises on all-NaT groups, so split processing into with-month and no-month groups.
    with_month = ordered.loc[ordered["month"].notna()].copy()
    latest_parts = []

    if not with_month.empty:
        latest_idx = with_month.groupby(keys)["month"].idxmax()
        latest_parts.append(ordered.loc[latest_idx].copy())

    no_month = ordered.loc[ordered["month"].isna()].copy()
    if not no_month.empty:
        fallback = no_month.drop_duplicates(subset=keys, keep="last")
        if latest_parts:
            existing_keys = latest_parts[0][keys].drop_duplicates()
            fallback = fallback.merge(existing_keys, on=keys, how="left", indicator=True)
            fallback = fallback.loc[fallback["_merge"] == "left_only"].drop(columns=["_merge"])
        latest_parts.append(fallback)

    latest = pd.concat(latest_parts, ignore_index=True) if latest_parts else ordered.head(0).copy()

    trends = (
        ordered.groupby(keys)
        .agg(
            min_cost=("cost_month", "min"),
            max_cost=("cost_month", "max"),
            min_rev=("rev_month", "min"),
            max_rev=("rev_month", "max"),
            months_observed=("month", "nunique"),
            avg_rev=("rev_month", "mean"),
            avg_cost=("cost_month", "mean"),
        )
        .reset_index()
    )

    snapshot = latest.merge(trends, on=keys, how="left")
    snapshot["fixed_bid_flag"] = snapshot["engagement_model"].str.upper().eq("FP")
    snapshot["cost_increased"] = snapshot["fixed_bid_flag"] & (snapshot["max_cost"] > snapshot["min_cost"])
    snapshot["revenue_changed"] = snapshot["fixed_bid_flag"] & (snapshot["max_rev"] != snapshot["min_rev"])

    latest_month = ordered["month"].max()
    snapshot["projected_margin"] = snapshot.apply(
        lambda row: _project_margin_for_row(row, latest_month),
        axis=1,
    )

    return snapshot


def _project_margin_for_row(row: pd.Series, latest_month: pd.Timestamp) -> float:
    start = row.get("start_date")
    end = row.get("end_date")
    if pd.isna(end):
        return float(row.get("margin_jtd", 0.0))

    # Use elapsed monthly run-rate from available months to project until the project end date.
    months_observed = int(row.get("months_observed", 0) or 0)
    if months_observed <= 0:
        return float(row.get("margin_jtd", 0.0))

    if pd.isna(start):
        start = latest_month

    total_months = max(1, (end.year - start.year) * 12 + (end.month - start.month) + 1)
    avg_margin = float(row.get("avg_rev", 0.0) - row.get("avg_cost", 0.0))
    return avg_margin * total_months


def summarize(data: pd.DataFrame, level: str) -> pd.DataFrame:
    if data.empty:
        return pd.DataFrame()

    keys = ["customer_name"] if level == "Account" else ["customer_name", "project_name"]
    grouped = (
        data.groupby(keys + ["month"], dropna=False)
        .agg(
            revenue_month=("rev_month", "sum"),
            cost_month=("cost_month", "sum"),
            revenue_ytd=("rev_ytd", "sum"),
            cost_ytd=("cost_ytd", "sum"),
            revenue_jtd=("rev_jtd", "sum"),
            cost_jtd=("cost_jtd", "sum"),
        )
        .reset_index()
    )

    grouped["margin_month"] = grouped["revenue_month"] - grouped["cost_month"]
    grouped["gm_month_pct"] = _safe_gm_pct(grouped["margin_month"], grouped["revenue_month"])

    grouped["margin_ytd"] = grouped["revenue_ytd"] - grouped["cost_ytd"]
    grouped["gm_ytd_pct"] = _safe_gm_pct(grouped["margin_ytd"], grouped["revenue_ytd"])

    grouped["margin_jtd"] = grouped["revenue_jtd"] - grouped["cost_jtd"]
    grouped["gm_jtd_pct"] = _safe_gm_pct(grouped["margin_jtd"], grouped["revenue_jtd"])

    grouped = grouped.sort_values(["customer_name", "month"])
    return grouped
